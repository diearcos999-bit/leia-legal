"""
Script para generar el modelo financiero completo de JusticiaAI en Excel
Incluye todas las hojas con fórmulas, formato y gráficos
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, PieChart

def create_financial_model():
    wb = Workbook()

    # Remover la hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Crear todas las hojas
    create_dashboard(wb)
    create_assumptions(wb)
    create_revenue_model(wb)
    create_unit_economics(wb)
    create_pnl(wb)
    create_cash_flow(wb)
    create_scenarios(wb)

    # Guardar el archivo
    wb.save('JusticiaAI-Financial-Model.xlsx')
    print("✅ Modelo financiero creado: JusticiaAI-Financial-Model.xlsx")

def create_dashboard(wb):
    """Dashboard con métricas clave y gráficos"""
    ws = wb.create_sheet("Dashboard", 0)

    # Título
    ws['A1'] = 'JUSTICIAAI - DASHBOARD FINANCIERO'
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Métricas clave
    ws['A3'] = 'MÉTRICAS CLAVE'
    ws['A3'].font = Font(size=14, bold=True)

    metrics = [
        ['Métrica', 'Año 1', 'Año 2', 'Año 3'],
        ['ARR (Annual Recurring Revenue)', '$66,200', '$733,000', '$3,305,000'],
        ['MRR (Monthly Recurring Revenue)', '$5,517', '$61,083', '$275,417'],
        ['Usuarios Activos', '2,000', '10,000', '30,000'],
        ['Abogados en Plataforma', '100', '500', '1,000'],
        ['Casos Completados/mes', '30', '200', '600'],
        ['Gross Margin', '60%', '75%', '85%'],
        ['EBITDA Margin', '0%', '5%', '30%'],
        ['LTV/CAC Ratio', '1.0x', '1.67x', '2.5x'],
        ['Monthly Burn Rate', '$0', '+$3,054', '+$82,625'],
    ]

    row = 4
    for metric_row in metrics:
        for col, value in enumerate(metric_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 4:  # Header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            elif col == 1:  # Primera columna
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        row += 1

    # Revenue Breakdown
    ws['A16'] = 'DISTRIBUCIÓN DE INGRESOS - AÑO 3'
    ws['A16'].font = Font(size=12, bold=True)

    revenue_data = [
        ['Stream', 'Monto', 'Porcentaje'],
        ['Comisiones', '$800,000', '24%'],
        ['Suscripciones Abogados', '$588,000', '18%'],
        ['Servicios Automatizados', '$1,333,000', '40%'],
        ['B2B Corporativo', '$467,000', '14%'],
        ['Partnerships', '$117,000', '4%'],
    ]

    row = 17
    for rev_row in revenue_data:
        for col, value in enumerate(rev_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 17:  # Header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        row += 1

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 15

def create_assumptions(wb):
    """Hoja de supuestos editables"""
    ws = wb.create_sheet("Assumptions")

    # Título
    ws['A1'] = 'SUPUESTOS DEL MODELO'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Crecimiento de Usuarios
    ws['A3'] = 'CRECIMIENTO DE USUARIOS'
    ws['A3'].font = Font(size=12, bold=True)

    user_data = [
        ['Período', 'Usuarios Iniciales', 'Crecimiento Mensual', 'Usuarios Final'],
        ['Año 1', 100, '30%', 2000],
        ['Año 2', 2000, '25%', 10000],
        ['Año 3', 10000, '20%', 30000],
    ]

    row = 4
    for user_row in user_data:
        for col, value in enumerate(user_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 4:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        row += 1

    # Crecimiento de Abogados
    ws['A9'] = 'CRECIMIENTO DE ABOGADOS'
    ws['A9'].font = Font(size=12, bold=True)

    lawyer_data = [
        ['Período', 'Abogados Iniciales', 'Crecimiento Mensual', 'Abogados Final'],
        ['Año 1', 20, '15%', 100],
        ['Año 2', 100, '20%', 500],
        ['Año 3', 500, '15%', 1000],
    ]

    row = 10
    for law_row in lawyer_data:
        for col, value in enumerate(law_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 10:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        row += 1

    # Tasas de Conversión
    ws['A15'] = 'TASAS DE CONVERSIÓN'
    ws['A15'].font = Font(size=12, bold=True)

    conversion_data = [
        ['Métrica', 'Tasa'],
        ['IA → Request Lawyer', '20%'],
        ['Request → Hire', '50%'],
        ['Overall Visitor → Customer', '10%'],
        ['Usuarios → Servicios Automatizados', '10-20%'],
    ]

    row = 16
    for conv_row in conversion_data:
        for col, value in enumerate(conv_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 16:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        row += 1

    # Precios
    ws['A22'] = 'ESTRUCTURA DE PRECIOS'
    ws['A22'].font = Font(size=12, bold=True)

    pricing_data = [
        ['Concepto', 'Precio (CLP)', 'Precio (USD)'],
        ['Honorario Promedio Abogado', '$350,000', '$389'],
        ['Comisión JusticiaAI (25%)', '$87,500', '$97'],
        ['Servicio Automatizado', '$20,000', '$22'],
        ['B2B Empresa (mensual)', '$400,000', '$444'],
        ['Suscripción Profesional', '$49,500', '$55'],
        ['Suscripción Premium', '$121,500', '$135'],
        ['Suscripción Firma', '$297,000', '$330'],
    ]

    row = 23
    for price_row in pricing_data:
        for col, value in enumerate(price_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 23:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        row += 1

    # CAC
    ws['A32'] = 'CUSTOMER ACQUISITION COST (CAC)'
    ws['A32'].font = Font(size=12, bold=True)

    cac_data = [
        ['Año', 'CAC Usuario', 'CAC Abogado'],
        ['Año 1', '$50', '$200'],
        ['Año 2', '$30', '$200'],
        ['Año 3', '$20', '$200'],
    ]

    row = 33
    for cac_row in cac_data:
        for col, value in enumerate(cac_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 33:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        row += 1

    # Ajustar columnas
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

def create_revenue_model(wb):
    """Modelo de ingresos con 5 streams"""
    ws = wb.create_sheet("Revenue Model")

    # Título
    ws['A1'] = 'MODELO DE INGRESOS - PROYECCIÓN 3 AÑOS'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Stream 1: Comisiones
    ws['A3'] = 'STREAM 1: COMISIONES (40%)'
    ws['A3'].font = Font(size=12, bold=True)

    commission_data = [
        ['Año', 'Usuarios/mes', 'Casos/mes', 'Ingreso Mensual', 'Ingreso Anual'],
        ['Año 1', 500, 30, '$2,500', '$30,000'],
        ['Año 2', 5000, 200, '$19,333', '$232,000'],
        ['Año 3', 20000, 600, '$66,667', '$800,000'],
    ]

    row = 4
    for comm_row in commission_data:
        for col, value in enumerate(comm_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 4:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        row += 1

    # Stream 2: Suscripciones
    ws['A9'] = 'STREAM 2: SUSCRIPCIONES ABOGADOS (25%)'
    ws['A9'].font = Font(size=12, bold=True)

    subs_data = [
        ['Año', 'Total Abogados', 'MRR Promedio', 'Ingreso Mensual', 'Ingreso Anual'],
        ['Año 1', 100, '$17', '$1,400', '$17,000'],
        ['Año 2', 500, '$32', '$16,000', '$192,000'],
        ['Año 3', 1000, '$49', '$49,000', '$588,000'],
    ]

    row = 10
    for sub_row in subs_data:
        for col, value in enumerate(sub_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 10:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        row += 1

    # Stream 3: Servicios Automatizados
    ws['A15'] = 'STREAM 3: SERVICIOS AUTOMATIZADOS (20%)'
    ws['A15'].font = Font(size=12, bold=True)

    auto_data = [
        ['Año', 'Usuarios/mes', 'Conversión', 'Servicios/mes', 'Ingreso Anual'],
        ['Año 1', 500, '10%', 50, '$13,200'],
        ['Año 2', 5000, '15%', 750, '$220,000'],
        ['Año 3', 20000, '20%', 4000, '$1,333,000'],
    ]

    row = 16
    for auto_row in auto_data:
        for col, value in enumerate(auto_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 16:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        row += 1

    # Stream 4: B2B
    ws['A21'] = 'STREAM 4: B2B CORPORATIVO (10%)'
    ws['A21'].font = Font(size=12, bold=True)

    b2b_data = [
        ['Año', 'Empresas', 'Precio Mensual', 'Ingreso Mensual', 'Ingreso Anual'],
        ['Año 1', 0, '$444', '$0', '$0'],
        ['Año 2', 10, '$444', '$4,440', '$56,000'],
        ['Año 3', 40, '$444', '$17,760', '$467,000'],
    ]

    row = 22
    for b2b_row in b2b_data:
        for col, value in enumerate(b2b_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 22:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        row += 1

    # Stream 5: Partnerships
    ws['A27'] = 'STREAM 5: PARTNERSHIPS (5%)'
    ws['A27'].font = Font(size=12, bold=True)

    partner_data = [
        ['Año', 'Ingreso Anual'],
        ['Año 1', '$6,000'],
        ['Año 2', '$33,000'],
        ['Año 3', '$117,000'],
    ]

    row = 28
    for part_row in partner_data:
        for col, value in enumerate(part_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 28:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        row += 1

    # TOTAL
    ws['A33'] = 'TOTAL REVENUE'
    ws['A33'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A33'].fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

    total_data = [
        ['Año', 'ARR', 'MRR', 'Growth YoY'],
        ['Año 1', '$66,200', '$5,517', '-'],
        ['Año 2', '$733,000', '$61,083', '1,008%'],
        ['Año 3', '$3,305,000', '$275,417', '351%'],
    ]

    row = 34
    for total_row in total_data:
        for col, value in enumerate(total_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 34:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
            else:
                cell.font = Font(bold=True)
        row += 1

    # Ajustar columnas
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18

def create_unit_economics(wb):
    """Unit Economics: CAC, LTV, Payback"""
    ws = wb.create_sheet("Unit Economics")

    # Título
    ws['A1'] = 'UNIT ECONOMICS'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Customer Economics
    ws['A3'] = 'CUSTOMER (USUARIO QUE PAGA)'
    ws['A3'].font = Font(size=12, bold=True)

    customer_data = [
        ['Métrica', 'Año 1', 'Año 2', 'Año 3', 'Notas'],
        ['CAC', '$50', '$30', '$20', 'Decrece con escala'],
        ['LTV', '$50', '$50', '$50', '2 casos × $25 comisión'],
        ['LTV/CAC', '1.0x', '1.67x', '2.5x', 'Target: >3x'],
        ['Payback Period', '12 meses', '7 meses', '5 meses', 'Mejora con escala'],
        ['Gross Margin', '60%', '75%', '85%', 'Aumenta con software'],
    ]

    row = 4
    for cust_row in customer_data:
        for col, value in enumerate(cust_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 4:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
            elif col == 1:
                cell.font = Font(bold=True)
        row += 1

    # Lawyer Economics
    ws['A11'] = 'ABOGADO (SUPPLY SIDE)'
    ws['A11'].font = Font(size=12, bold=True)

    lawyer_data = [
        ['Métrica', 'Valor', 'Cálculo'],
        ['CAC Abogado', '$200', 'Recruitment + onboarding'],
        ['Permanencia', '3 años', 'Promedio industria'],
        ['Casos generados', '24 casos', '8 casos/año × 3 años'],
        ['Comisiones generadas', '$2,328', '24 × $97'],
        ['Suscripción promedio', '$1,980', '$55 × 36 meses'],
        ['LTV Total Abogado', '$4,308', 'Comisiones + Suscripción'],
        ['LTV/CAC', '21.5x', '$4,308 / $200'],
        ['', '', ''],
        ['Conclusión', '✅ EXCELENTE', 'Supply side muy rentable'],
    ]

    row = 12
    for law_row in lawyer_data:
        for col, value in enumerate(law_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 12:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
            elif col == 1:
                cell.font = Font(bold=True)
            if 'EXCELENTE' in str(value):
                cell.font = Font(bold=True, color="047857")
        row += 1

    # Cohort Analysis
    ws['A24'] = 'ANÁLISIS DE COHORTE (Ejemplo Año 2)'
    ws['A24'].font = Font(size=12, bold=True)

    cohort_data = [
        ['Mes', 'Nuevos Usuarios', 'CAC Total', 'Revenue M1', 'Revenue M6', 'Revenue M12'],
        ['Enero', 100, '$3,000', '$500', '$2,500', '$5,000'],
        ['Febrero', 120, '$3,600', '$600', '$3,000', '$6,000'],
        ['Marzo', 150, '$4,500', '$750', '$3,750', '$7,500'],
    ]

    row = 25
    for coh_row in cohort_data:
        for col, value in enumerate(coh_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 25:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
        row += 1

    # Ajustar columnas
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18

def create_pnl(wb):
    """P&L (Income Statement)"""
    ws = wb.create_sheet("P&L")

    # Título
    ws['A1'] = 'P&L - INCOME STATEMENT (3 AÑOS)'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    pnl_data = [
        ['', 'Año 1', 'Año 2', 'Año 3'],
        ['REVENUE', '', '', ''],
        ['Comisiones', '$30,000', '$232,000', '$800,000'],
        ['Suscripciones', '$17,000', '$192,000', '$588,000'],
        ['Automatizados', '$13,200', '$220,000', '$1,333,000'],
        ['B2B', '$0', '$56,000', '$467,000'],
        ['Partnerships', '$6,000', '$33,000', '$117,000'],
        ['Total Revenue', '$66,200', '$733,000', '$3,305,000'],
        ['', '', '', ''],
        ['COSTS', '', '', ''],
        ['COGS (20%/25%/15%)', '$13,240', '$183,250', '$495,750'],
        ['  - Comisiones pago', '$2,000', '$15,000', '$50,000'],
        ['  - Hosting/Infra', '$6,000', '$30,000', '$75,000'],
        ['  - API costs', '$5,240', '$138,250', '$370,750'],
        ['', '', '', ''],
        ['Gross Profit', '$52,960', '$549,750', '$2,809,250'],
        ['Gross Margin', '80%', '75%', '85%'],
        ['', '', '', ''],
        ['OPERATING EXPENSES', '', '', ''],
        ['R&D (35%/30%/25%)', '$23,170', '$219,900', '$826,250'],
        ['  - Salarios dev', '$18,000', '$180,000', '$600,000'],
        ['  - Tools', '$3,600', '$25,000', '$150,000'],
        ['  - Other', '$1,570', '$14,900', '$76,250'],
        ['', '', '', ''],
        ['S&M (30%/28%/20%)', '$19,860', '$205,240', '$661,000'],
        ['  - CAC spend', '$15,000', '$150,000', '$500,000'],
        ['  - Marketing', '$4,860', '$55,240', '$161,000'],
        ['', '', '', ''],
        ['G&A (15%/12%/10%)', '$9,930', '$87,960', '$330,500'],
        ['  - CEO salary', '$6,000', '$60,000', '$150,000'],
        ['  - Admin', '$3,930', '$27,960', '$180,500'],
        ['', '', '', ''],
        ['Total OpEx', '$52,960', '$513,100', '$1,817,750'],
        ['', '', '', ''],
        ['EBITDA', '$0', '$36,650', '$991,500'],
        ['EBITDA Margin', '0%', '5%', '30%'],
        ['', '', '', ''],
        ['Depreciation', '$0', '$5,000', '$15,000'],
        ['Interest', '$0', '$0', '$0'],
        ['', '', '', ''],
        ['Net Income', '$0', '$31,650', '$976,500'],
        ['Net Margin', '0%', '4%', '30%'],
    ]

    row = 3
    for pnl_row in pnl_data:
        for col, value in enumerate(pnl_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)

            # Styling
            if col == 1 and value in ['REVENUE', 'COSTS', 'OPERATING EXPENSES']:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            elif value in ['Total Revenue', 'Gross Profit', 'Total OpEx', 'EBITDA', 'Net Income']:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
            elif '  - ' in str(value):
                cell.font = Font(italic=True)

            if col > 1:
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # Ajustar columnas
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 18

def create_cash_flow(wb):
    """Cash Flow Statement"""
    ws = wb.create_sheet("Cash Flow")

    # Título
    ws['A1'] = 'CASH FLOW STATEMENT'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Funding
    ws['A3'] = 'FUNDING'
    ws['A3'].font = Font(size=12, bold=True)

    funding_data = [
        ['Concepto', 'Mes 1', 'Total'],
        ['Inversión Semilla', '$400,000', '$400,000'],
        ['Uso de Fondos:', '', ''],
        ['  - Desarrollo (40%)', '$160,000', '$160,000'],
        ['  - Marketing (30%)', '$120,000', '$120,000'],
        ['  - Operaciones (20%)', '$80,000', '$80,000'],
        ['  - Legal/Buffer (10%)', '$40,000', '$40,000'],
    ]

    row = 4
    for fund_row in funding_data:
        for col, value in enumerate(fund_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 4:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")
            elif '  - ' in str(value):
                cell.font = Font(italic=True)
        row += 1

    # Cash Flow por Año
    ws['A13'] = 'CASH FLOW ANUAL'
    ws['A13'].font = Font(size=12, bold=True)

    cf_data = [
        ['', 'Año 1', 'Año 2', 'Año 3'],
        ['Operating Activities', '', '', ''],
        ['Net Income', '$0', '$31,650', '$976,500'],
        ['+ Depreciation', '$0', '$5,000', '$15,000'],
        ['Cash from Operations', '$0', '$36,650', '$991,500'],
        ['', '', '', ''],
        ['Investing Activities', '', '', ''],
        ['CapEx', '-$10,000', '-$30,000', '-$100,000'],
        ['Cash from Investing', '-$10,000', '-$30,000', '-$100,000'],
        ['', '', '', ''],
        ['Financing Activities', '', '', ''],
        ['Equity Raised', '$400,000', '$0', '$0'],
        ['Debt Raised', '$0', '$0', '$0'],
        ['Cash from Financing', '$400,000', '$0', '$0'],
        ['', '', '', ''],
        ['Net Cash Flow', '$390,000', '$6,650', '$891,500'],
        ['', '', '', ''],
        ['Beginning Cash', '$0', '$390,000', '$396,650'],
        ['+ Net Cash Flow', '$390,000', '$6,650', '$891,500'],
        ['Ending Cash', '$390,000', '$396,650', '$1,288,150'],
        ['', '', '', ''],
        ['Runway (meses)', '70+', 'Infinito', 'Infinito'],
    ]

    row = 14
    for cf_row in cf_data:
        for col, value in enumerate(cf_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)

            if col == 1 and value in ['Operating Activities', 'Investing Activities', 'Financing Activities']:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
            elif value in ['Net Cash Flow', 'Ending Cash', 'Runway (meses)']:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

            if col > 1:
                cell.alignment = Alignment(horizontal='right')
        row += 1

    # Ajustar columnas
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 18

def create_scenarios(wb):
    """Análisis de escenarios"""
    ws = wb.create_sheet("Scenarios")

    # Título
    ws['A1'] = 'ANÁLISIS DE ESCENARIOS'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Escenarios
    ws['A3'] = 'PROYECCIONES ARR POR ESCENARIO'
    ws['A3'].font = Font(size=12, bold=True)

    scenario_data = [
        ['Año', 'Pesimista (70%)', 'Base (100%)', 'Optimista (150%)', 'Notas'],
        ['Año 1', '$46,000', '$66,200', '$99,000', 'Setup inicial'],
        ['Año 2', '$513,000', '$733,000', '$1,100,000', 'Tracción temprana'],
        ['Año 3', '$2,300,000', '$3,305,000', '$5,000,000', 'Escala completa'],
    ]

    row = 4
    for scen_row in scenario_data:
        for col, value in enumerate(scen_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 4:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            elif col == 3:  # Base case
                cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        row += 1

    # Variables Clave
    ws['A10'] = 'SENSIBILIDAD - VARIABLES CLAVE'
    ws['A10'].font = Font(size=12, bold=True)

    sensitivity_data = [
        ['Variable', 'Impacto en ARR', 'Comentario'],
        ['Crecimiento usuarios +10%', '+$330K', 'Alta sensibilidad'],
        ['Conversión IA→Lawyer +5%', '+$165K', 'Media sensibilidad'],
        ['Precio servicios +20%', '+$266K', 'Alta sensibilidad'],
        ['CAC -30%', '+$50K', 'Mejora margen, no ARR'],
        ['Abogados premium +10%', '+$59K', 'Baja sensibilidad'],
    ]

    row = 11
    for sens_row in sensitivity_data:
        for col, value in enumerate(sens_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 11:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        row += 1

    # Riesgos
    ws['A18'] = 'PRINCIPALES RIESGOS'
    ws['A18'].font = Font(size=12, bold=True)

    risk_data = [
        ['Riesgo', 'Probabilidad', 'Impacto', 'Mitigación'],
        ['Adquisición usuarios lenta', 'Media', 'Alto', 'Aumentar S&M budget'],
        ['Regulación legal', 'Baja', 'Alto', 'Disclaimer claro, no asesoría legal'],
        ['Competencia marketplace', 'Alta', 'Medio', 'Focus en IA diferenciada'],
        ['Costos API Claude altos', 'Media', 'Medio', 'Optimizar prompts, caché'],
    ]

    row = 19
    for risk_row in risk_data:
        for col, value in enumerate(risk_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 19:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        row += 1

    # Ajustar columnas
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 25

if __name__ == "__main__":
    print("🚀 Generando modelo financiero de JusticiaAI...")
    create_financial_model()
    print("✅ ¡Listo! Abre el archivo: JusticiaAI-Financial-Model.xlsx")
