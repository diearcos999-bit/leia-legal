"""
Script para generar el modelo financiero AVANZADO de JusticiaAI en Excel
Con FÓRMULAS DINÁMICAS que vinculan todas las hojas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, PieChart

def create_financial_model_advanced():
    wb = Workbook()

    # Remover la hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Crear hojas en orden (Assumptions primero porque otras hojas referencian a esta)
    create_assumptions_advanced(wb)
    create_revenue_model_advanced(wb)
    create_unit_economics_advanced(wb)
    create_pnl_advanced(wb)
    create_cash_flow_advanced(wb)
    create_scenarios_advanced(wb)
    create_dashboard_advanced(wb)  # Dashboard al final para referenciar todo

    # Guardar el archivo
    wb.save('JusticiaAI-Financial-Model-Advanced.xlsx')
    print("✅ Modelo financiero AVANZADO creado: JusticiaAI-Financial-Model-Advanced.xlsx")
    print("🔗 Todas las hojas están vinculadas con fórmulas dinámicas")

def create_assumptions_advanced(wb):
    """Hoja de supuestos con valores base que se pueden editar"""
    ws = wb.create_sheet("Assumptions", 0)

    # Título
    ws['A1'] = 'SUPUESTOS DEL MODELO (EDITA ESTOS VALORES)'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25

    # Nota importante
    ws['A2'] = 'IMPORTANTE: Cambia estos valores y todas las hojas se actualizarán automáticamente'
    ws['A2'].font = Font(italic=True, color="DC2626")
    ws.merge_cells('A2:E2')

    # USUARIOS
    ws['A4'] = 'CRECIMIENTO DE USUARIOS'
    ws['A4'].font = Font(size=12, bold=True)
    ws['A4'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws.merge_cells('A4:E4')

    headers = ['Métrica', 'Año 1', 'Año 2', 'Año 3', 'Unidad']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    user_assumptions = [
        ['Usuarios iniciales', 100, 2000, 10000, 'usuarios'],
        ['Crecimiento mensual', 0.30, 0.25, 0.20, '%'],
        ['Usuarios finales (promedio)', 2000, 10000, 30000, 'usuarios'],
    ]

    row = 6
    for ass_row in user_assumptions:
        for col, value in enumerate(ass_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # ABOGADOS
    ws['A10'] = 'CRECIMIENTO DE ABOGADOS'
    ws['A10'].font = Font(size=12, bold=True)
    ws['A10'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws.merge_cells('A10:E10')

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    lawyer_assumptions = [
        ['Abogados iniciales', 20, 100, 500, 'abogados'],
        ['Crecimiento mensual', 0.15, 0.20, 0.15, '%'],
        ['Abogados finales (promedio)', 100, 500, 1000, 'abogados'],
    ]

    row = 12
    for ass_row in lawyer_assumptions:
        for col, value in enumerate(ass_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # CONVERSIONES
    ws['A16'] = 'TASAS DE CONVERSIÓN'
    ws['A16'].font = Font(size=12, bold=True)
    ws['A16'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    ws.merge_cells('A16:E16')

    conv_headers = ['Métrica', 'Tasa', 'Descripción', '', '']
    for col, header in enumerate(conv_headers, start=1):
        cell = ws.cell(row=17, column=col, value=header)
        if header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    conversion_data = [
        ['IA → Request Lawyer', 0.20, 'Usuarios que piden abogado'],
        ['Request → Hire', 0.50, 'Requests que contratan'],
        ['Overall Conversion', 0.10, 'Usuario → Cliente pagador'],
        ['Usuarios → Servicios Auto Y1', 0.10, 'Conversión servicios automatizados'],
        ['Usuarios → Servicios Auto Y2', 0.15, ''],
        ['Usuarios → Servicios Auto Y3', 0.20, ''],
    ]

    row = 18
    for conv_row in conversion_data:
        for col, value in enumerate(conv_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # PRECIOS
    ws['A26'] = 'PRECIOS (USD)'
    ws['A26'].font = Font(size=12, bold=True)
    ws['A26'].fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
    ws.merge_cells('A26:E26')

    price_headers = ['Concepto', 'Precio USD', 'Notas', '', '']
    for col, header in enumerate(price_headers, start=1):
        cell = ws.cell(row=27, column=col, value=header)
        if header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid")

    pricing_data = [
        ['Honorario promedio abogado', 389, 'CLP $350,000'],
        ['Comisión JusticiaAI', 0.25, '25% de honorario'],
        ['Ingreso por caso', 97, '$389 × 25%'],
        ['Servicio automatizado', 22, 'CLP $20,000'],
        ['B2B empresa/mes', 444, 'CLP $400,000'],
        ['Suscripción Free', 0, ''],
        ['Suscripción Profesional', 55, 'CLP $49,500'],
        ['Suscripción Premium', 135, 'CLP $121,500'],
        ['Suscripción Firma', 330, 'CLP $297,000'],
    ]

    row = 28
    for price_row in pricing_data:
        for col, value in enumerate(price_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # MIX DE SUSCRIPCIONES
    ws['A39'] = 'MIX DE SUSCRIPCIONES ABOGADOS'
    ws['A39'].font = Font(size=12, bold=True)
    ws['A39'].fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    ws.merge_cells('A39:E39')

    sub_headers = ['Plan', 'Año 1', 'Año 2', 'Año 3', 'Precio']
    for col, header in enumerate(sub_headers, start=1):
        cell = ws.cell(row=40, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")

    subscription_mix = [
        ['Free', 0.60, 0.50, 0.40, 0],
        ['Profesional', 0.30, 0.35, 0.45, 55],
        ['Premium', 0.10, 0.13, 0.13, 135],
        ['Firma', 0.00, 0.02, 0.02, 330],
    ]

    row = 41
    for sub_row in subscription_mix:
        for col, value in enumerate(sub_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # CAC
    ws['A47'] = 'CUSTOMER ACQUISITION COST (CAC)'
    ws['A47'].font = Font(size=12, bold=True)
    ws['A47'].fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
    ws.merge_cells('A47:E47')

    cac_headers = ['Tipo', 'Año 1', 'Año 2', 'Año 3', 'Unidad']
    for col, header in enumerate(cac_headers, start=1):
        cell = ws.cell(row=48, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    cac_data = [
        ['CAC Usuario', 50, 30, 20, 'USD'],
        ['CAC Abogado', 200, 200, 200, 'USD'],
    ]

    row = 49
    for cac_row in cac_data:
        for col, value in enumerate(cac_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # B2B EMPRESAS
    ws['A53'] = 'B2B CORPORATIVO'
    ws['A53'].font = Font(size=12, bold=True)
    ws['A53'].fill = PatternFill(start_color="06B6D4", end_color="06B6D4", fill_type="solid")
    ws.merge_cells('A53:E53')

    b2b_headers = ['Métrica', 'Año 1', 'Año 2', 'Año 3', '']
    for col, header in enumerate(b2b_headers, start=1):
        cell = ws.cell(row=54, column=col, value=header)
        if header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CFFAFE", end_color="CFFAFE", fill_type="solid")

    b2b_data = [
        ['Empresas clientes', 0, 10, 40, ''],
        ['Precio mensual', 444, 444, 444, 'USD'],
    ]

    row = 55
    for b2b_row in b2b_data:
        for col, value in enumerate(b2b_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # COSTOS OPERATIVOS
    ws['A59'] = 'ESTRUCTURA DE COSTOS (%)'
    ws['A59'].font = Font(size=12, bold=True)
    ws['A59'].fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")
    ws.merge_cells('A59:E59')

    cost_headers = ['Categoría', 'Año 1', 'Año 2', 'Año 3', 'Notas']
    for col, header in enumerate(cost_headers, start=1):
        cell = ws.cell(row=60, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    cost_structure = [
        ['COGS % Revenue', 0.20, 0.25, 0.15, 'Disminuye con escala'],
        ['R&D % Revenue', 0.35, 0.30, 0.25, 'Desarrollo + tools'],
        ['S&M % Revenue', 0.30, 0.28, 0.20, 'Marketing + CAC'],
        ['G&A % Revenue', 0.15, 0.12, 0.10, 'Admin + CEO'],
    ]

    row = 61
    for cost_row in cost_structure:
        for col, value in enumerate(cost_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # FUNDING
    ws['A67'] = 'FUNDING'
    ws['A67'].font = Font(size=12, bold=True)
    ws['A67'].fill = PatternFill(start_color="14B8A6", end_color="14B8A6", fill_type="solid")
    ws.merge_cells('A67:E67')

    fund_headers = ['Concepto', 'Monto USD', '', '', '']
    for col, header in enumerate(fund_headers, start=1):
        cell = ws.cell(row=68, column=col, value=header)
        if header:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")

    funding_data = [
        ['Inversión Semilla', 400000],
        ['% Desarrollo', 0.40],
        ['% Marketing', 0.30],
        ['% Operaciones', 0.20],
        ['% Legal/Buffer', 0.10],
    ]

    row = 69
    for fund_row in funding_data:
        for col, value in enumerate(fund_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18

def create_revenue_model_advanced(wb):
    """Revenue Model con fórmulas que referencian Assumptions"""
    ws = wb.create_sheet("Revenue Model")

    # Título
    ws['A1'] = 'MODELO DE INGRESOS - PROYECCIÓN 3 AÑOS'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers de años
    ws['B3'] = 'Año 1'
    ws['C3'] = 'Año 2'
    ws['D3'] = 'Año 3'
    ws['E3'] = 'Fórmula'
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        ws[f'{col}3'].alignment = Alignment(horizontal='center')

    # STREAM 1: COMISIONES
    ws['A5'] = 'STREAM 1: COMISIONES'
    ws['A5'].font = Font(size=12, bold=True)

    ws['A6'] = 'Usuarios promedio/mes'
    ws['B6'] = '=Assumptions!C8'  # Año 1 usuarios finales
    ws['C6'] = '=Assumptions!D8'
    ws['D6'] = '=Assumptions!E8'

    ws['A7'] = 'Conversión global'
    ws['B7'] = '=Assumptions!B20'  # Overall conversion
    ws['C7'] = '=Assumptions!B20'
    ws['D7'] = '=Assumptions!B20'

    ws['A8'] = 'Casos/mes'
    ws['B8'] = '=B6*B7'
    ws['C8'] = '=C6*C7'
    ws['D8'] = '=D6*D7'

    ws['A9'] = 'Ingreso por caso'
    ws['B9'] = '=Assumptions!B30'  # Ingreso por caso
    ws['C9'] = '=Assumptions!B30'
    ws['D9'] = '=Assumptions!B30'

    ws['A10'] = 'Ingreso mensual'
    ws['B10'] = '=B8*B9'
    ws['C10'] = '=C8*C9'
    ws['D10'] = '=D8*D9'

    ws['A11'] = 'INGRESO ANUAL COMISIONES'
    ws['B11'] = '=B10*12'
    ws['C11'] = '=C10*12'
    ws['D11'] = '=D10*12'
    ws['A11'].font = Font(bold=True)
    ws['A11'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}11'].font = Font(bold=True)
        ws[f'{col}11'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    # STREAM 2: SUSCRIPCIONES
    ws['A14'] = 'STREAM 2: SUSCRIPCIONES ABOGADOS'
    ws['A14'].font = Font(size=12, bold=True)

    ws['A15'] = 'Total abogados'
    ws['B15'] = '=Assumptions!C14'
    ws['C15'] = '=Assumptions!D14'
    ws['D15'] = '=Assumptions!E14'

    ws['A16'] = '% Free (0 USD)'
    ws['B16'] = '=Assumptions!B41'
    ws['C16'] = '=Assumptions!C41'
    ws['D16'] = '=Assumptions!D41'

    ws['A17'] = '% Profesional (55 USD)'
    ws['B17'] = '=Assumptions!B42'
    ws['C17'] = '=Assumptions!C42'
    ws['D17'] = '=Assumptions!D42'

    ws['A18'] = '% Premium (135 USD)'
    ws['B18'] = '=Assumptions!B43'
    ws['C18'] = '=Assumptions!C43'
    ws['D18'] = '=Assumptions!D43'

    ws['A19'] = '% Firma (330 USD)'
    ws['B19'] = '=Assumptions!B44'
    ws['C19'] = '=Assumptions!C44'
    ws['D19'] = '=Assumptions!D44'

    ws['A20'] = 'MRR Suscripciones'
    ws['B20'] = '=(B15*B17*Assumptions!E42)+(B15*B18*Assumptions!E43)+(B15*B19*Assumptions!E44)'
    ws['C20'] = '=(C15*C17*Assumptions!E42)+(C15*C18*Assumptions!E43)+(C15*C19*Assumptions!E44)'
    ws['D20'] = '=(D15*D17*Assumptions!E42)+(D15*D18*Assumptions!E43)+(D15*D19*Assumptions!E44)'

    ws['A21'] = 'INGRESO ANUAL SUSCRIPCIONES'
    ws['B21'] = '=B20*12'
    ws['C21'] = '=C20*12'
    ws['D21'] = '=D20*12'
    ws['A21'].font = Font(bold=True)
    ws['A21'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}21'].font = Font(bold=True)
        ws[f'{col}21'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    # STREAM 3: SERVICIOS AUTOMATIZADOS
    ws['A24'] = 'STREAM 3: SERVICIOS AUTOMATIZADOS'
    ws['A24'].font = Font(size=12, bold=True)

    ws['A25'] = 'Usuarios promedio/mes'
    ws['B25'] = '=Assumptions!C8'
    ws['C25'] = '=Assumptions!D8'
    ws['D25'] = '=Assumptions!E8'

    ws['A26'] = 'Conversión servicios'
    ws['B26'] = '=Assumptions!B21'
    ws['C26'] = '=Assumptions!B22'
    ws['D26'] = '=Assumptions!B23'

    ws['A27'] = 'Servicios vendidos/mes'
    ws['B27'] = '=B25*B26'
    ws['C27'] = '=C25*C26'
    ws['D27'] = '=D25*D26'

    ws['A28'] = 'Precio servicio'
    ws['B28'] = '=Assumptions!B31'
    ws['C28'] = '=Assumptions!B31'
    ws['D28'] = '=Assumptions!B31'

    ws['A29'] = 'INGRESO ANUAL AUTOMATIZADOS'
    ws['B29'] = '=B27*B28*12'
    ws['C29'] = '=C27*C28*12'
    ws['D29'] = '=D27*D28*12'
    ws['A29'].font = Font(bold=True)
    ws['A29'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}29'].font = Font(bold=True)
        ws[f'{col}29'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    # STREAM 4: B2B
    ws['A32'] = 'STREAM 4: B2B CORPORATIVO'
    ws['A32'].font = Font(size=12, bold=True)

    ws['A33'] = 'Empresas clientes'
    ws['B33'] = '=Assumptions!B55'
    ws['C33'] = '=Assumptions!C55'
    ws['D33'] = '=Assumptions!D55'

    ws['A34'] = 'Precio mensual'
    ws['B34'] = '=Assumptions!B32'
    ws['C34'] = '=Assumptions!B32'
    ws['D34'] = '=Assumptions!B32'

    ws['A35'] = 'INGRESO ANUAL B2B'
    ws['B35'] = '=B33*B34*12'
    ws['C35'] = '=C33*C34*12'
    ws['D35'] = '=D33*D34*12'
    ws['A35'].font = Font(bold=True)
    ws['A35'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}35'].font = Font(bold=True)
        ws[f'{col}35'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    # STREAM 5: PARTNERSHIPS (simplificado, 5% del total)
    ws['A38'] = 'STREAM 5: PARTNERSHIPS'
    ws['A38'].font = Font(size=12, bold=True)

    ws['A39'] = 'INGRESO ANUAL PARTNERSHIPS'
    ws['B39'] = '=(B11+B21+B29+B35)*0.05'
    ws['C39'] = '=(C11+C21+C29+C35)*0.05'
    ws['D39'] = '=(D11+D21+D29+D35)*0.05'
    ws['A39'].font = Font(bold=True)
    ws['A39'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}39'].font = Font(bold=True)
        ws[f'{col}39'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    # TOTAL
    ws['A42'] = 'TOTAL REVENUE'
    ws['A42'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A42'].fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

    ws['A43'] = 'ARR (Annual Recurring Revenue)'
    ws['B43'] = '=B11+B21+B29+B35+B39'
    ws['C43'] = '=C11+C21+C29+C35+C39'
    ws['D43'] = '=D11+D21+D29+D35+D39'
    ws['A43'].font = Font(bold=True, size=12)
    for col in ['B', 'C', 'D']:
        ws[f'{col}43'].font = Font(bold=True, size=12)
        ws[f'{col}43'].number_format = '$#,##0'

    ws['A44'] = 'MRR (Monthly Recurring Revenue)'
    ws['B44'] = '=B43/12'
    ws['C44'] = '=C43/12'
    ws['D44'] = '=D43/12'
    ws['A44'].font = Font(bold=True)
    for col in ['B', 'C', 'D']:
        ws[f'{col}44'].font = Font(bold=True)
        ws[f'{col}44'].number_format = '$#,##0'

    ws['A45'] = 'Growth YoY'
    ws['B45'] = '-'
    ws['C45'] = '=(C43-B43)/B43'
    ws['D45'] = '=(D43-C43)/C43'
    ws['C45'].number_format = '0.0%'
    ws['D45'].number_format = '0.0%'

    # Formatear números
    for row in range(6, 40):
        for col in ['B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            if row in [11, 21, 29, 35, 39]:
                cell.number_format = '$#,##0'
            elif row in [8, 27, 33]:
                cell.number_format = '#,##0'
            elif row in [7, 16, 17, 18, 19, 26]:
                cell.number_format = '0%'
            elif row in [9, 10, 20, 28, 34]:
                cell.number_format = '$#,##0'

    # Ajustar columnas
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18

def create_unit_economics_advanced(wb):
    """Unit Economics con fórmulas"""
    ws = wb.create_sheet("Unit Economics")

    # Título
    ws['A1'] = 'UNIT ECONOMICS'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    ws['B3'] = 'Año 1'
    ws['C3'] = 'Año 2'
    ws['D3'] = 'Año 3'
    ws['E3'] = 'Target'
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")

    # CUSTOMER ECONOMICS
    ws['A5'] = 'CUSTOMER (USUARIO QUE PAGA)'
    ws['A5'].font = Font(size=12, bold=True)

    ws['A6'] = 'CAC (Customer Acquisition Cost)'
    ws['B6'] = '=Assumptions!B49'
    ws['C6'] = '=Assumptions!C49'
    ws['D6'] = '=Assumptions!D49'
    ws['E6'] = '< $30'

    ws['A7'] = 'LTV (Lifetime Value)'
    ws['B7'] = 50  # Valor fijo estimado
    ws['C7'] = 50
    ws['D7'] = 50
    ws['E7'] = '> $150'

    ws['A8'] = 'LTV/CAC Ratio'
    ws['B8'] = '=B7/B6'
    ws['C8'] = '=C7/C6'
    ws['D8'] = '=D7/D6'
    ws['E8'] = '> 3.0x'

    ws['A9'] = 'Payback Period (meses)'
    ws['B9'] = 12
    ws['C9'] = 7
    ws['D9'] = 5
    ws['E9'] = '< 12'

    ws['A10'] = 'Gross Margin'
    ws['B10'] = '=1-Assumptions!B61'
    ws['C10'] = '=1-Assumptions!C61'
    ws['D10'] = '=1-Assumptions!D61'
    ws['E10'] = '> 70%'

    # Formatear
    for row in [6, 7]:
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].number_format = '$#,##0'
    for col in ['B', 'C', 'D']:
        ws[f'{col}8'].number_format = '0.0"x"'
        ws[f'{col}10'].number_format = '0%'

    # LAWYER ECONOMICS
    ws['A13'] = 'ABOGADO (SUPPLY SIDE)'
    ws['A13'].font = Font(size=12, bold=True)

    ws['A14'] = 'CAC Abogado'
    ws['B14'] = '=Assumptions!B50'
    ws['C14'] = '=Assumptions!C50'
    ws['D14'] = '=Assumptions!D50'
    ws['E14'] = '< $250'

    ws['A15'] = 'Permanencia (años)'
    ws['B15'] = 3
    ws['C15'] = 3
    ws['D15'] = 3
    ws['E15'] = '> 2'

    ws['A16'] = 'Casos generados (total)'
    ws['B16'] = '=B15*8'
    ws['C16'] = '=C15*8'
    ws['D16'] = '=D15*8'
    ws['E16'] = ''

    ws['A17'] = 'Comisiones generadas'
    ws['B17'] = '=B16*Assumptions!B30'
    ws['C17'] = '=C16*Assumptions!B30'
    ws['D17'] = '=D16*Assumptions!B30'
    ws['E17'] = ''

    ws['A18'] = 'Suscripción promedio/mes'
    ws['B18'] = 55
    ws['C18'] = 55
    ws['D18'] = 55
    ws['E18'] = ''

    ws['A19'] = 'Suscripciones totales'
    ws['B19'] = '=B18*B15*12'
    ws['C19'] = '=C18*C15*12'
    ws['D19'] = '=D18*D15*12'
    ws['E19'] = ''

    ws['A20'] = 'LTV Total Abogado'
    ws['B20'] = '=B17+B19'
    ws['C20'] = '=C17+C19'
    ws['D20'] = '=D17+D19'
    ws['E20'] = '> $3,000'

    ws['A21'] = 'LTV/CAC Abogado'
    ws['B21'] = '=B20/B14'
    ws['C21'] = '=C20/C14'
    ws['D21'] = '=D20/D14'
    ws['E21'] = '> 10x'

    # Formatear
    for row in [14, 17, 18, 19, 20]:
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].number_format = '$#,##0'
    for col in ['B', 'C', 'D']:
        ws[f'{col}21'].number_format = '0.0"x"'

    # Conclusión
    ws['A24'] = 'CONCLUSIÓN'
    ws['A24'].font = Font(size=12, bold=True, color="047857")
    ws['A24'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    ws.merge_cells('A24:E24')

    ws['A25'] = 'Supply side (abogados) es MUY rentable con LTV/CAC de 21.5x ✅'
    ws['A25'].font = Font(bold=True, color="047857")
    ws.merge_cells('A25:E25')

    ws['A26'] = 'Demand side (usuarios) mejora con escala, alcanzando 2.5x en Año 3 ✅'
    ws['A26'].font = Font(italic=True)
    ws.merge_cells('A26:E26')

    # Ajustar columnas
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18

def create_pnl_advanced(wb):
    """P&L con fórmulas que referencian Revenue Model y Assumptions"""
    ws = wb.create_sheet("P&L")

    # Título
    ws['A1'] = 'P&L - INCOME STATEMENT (3 AÑOS)'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    ws['B3'] = 'Año 1'
    ws['C3'] = 'Año 2'
    ws['D3'] = 'Año 3'
    for col in ['B', 'C', 'D']:
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
        ws[f'{col}3'].alignment = Alignment(horizontal='center')

    # REVENUE
    ws['A5'] = 'REVENUE'
    ws['A5'].font = Font(size=12, bold=True)

    ws['A6'] = 'Comisiones'
    ws['B6'] = "='Revenue Model'!B11"
    ws['C6'] = "='Revenue Model'!C11"
    ws['D6'] = "='Revenue Model'!D11"

    ws['A7'] = 'Suscripciones'
    ws['B7'] = "='Revenue Model'!B21"
    ws['C7'] = "='Revenue Model'!C21"
    ws['D7'] = "='Revenue Model'!D21"

    ws['A8'] = 'Automatizados'
    ws['B8'] = "='Revenue Model'!B29"
    ws['C8'] = "='Revenue Model'!C29"
    ws['D8'] = "='Revenue Model'!D29"

    ws['A9'] = 'B2B'
    ws['B9'] = "='Revenue Model'!B35"
    ws['C9'] = "='Revenue Model'!C35"
    ws['D9'] = "='Revenue Model'!D35"

    ws['A10'] = 'Partnerships'
    ws['B10'] = "='Revenue Model'!B39"
    ws['C10'] = "='Revenue Model'!C39"
    ws['D10'] = "='Revenue Model'!D39"

    ws['A12'] = 'Total Revenue'
    ws['B12'] = '=SUM(B6:B10)'
    ws['C12'] = '=SUM(C6:C10)'
    ws['D12'] = '=SUM(D6:D10)'
    ws['A12'].font = Font(bold=True)
    ws['A12'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}12'].font = Font(bold=True)
        ws[f'{col}12'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    # COGS
    ws['A15'] = 'COGS'
    ws['A15'].font = Font(size=12, bold=True)

    ws['A16'] = 'Total COGS'
    ws['B16'] = '=B12*Assumptions!B61'
    ws['C16'] = '=C12*Assumptions!C61'
    ws['D16'] = '=D12*Assumptions!D61'

    ws['A18'] = 'Gross Profit'
    ws['B18'] = '=B12-B16'
    ws['C18'] = '=C12-C16'
    ws['D18'] = '=D12-D16'
    ws['A18'].font = Font(bold=True)
    ws['A18'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}18'].font = Font(bold=True)
        ws[f'{col}18'].fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    ws['A19'] = 'Gross Margin'
    ws['B19'] = '=B18/B12'
    ws['C19'] = '=C18/C12'
    ws['D19'] = '=D18/D12'
    for col in ['B', 'C', 'D']:
        ws[f'{col}19'].number_format = '0%'

    # OPERATING EXPENSES
    ws['A22'] = 'OPERATING EXPENSES'
    ws['A22'].font = Font(size=12, bold=True)

    ws['A23'] = 'R&D'
    ws['B23'] = '=B12*Assumptions!B62'
    ws['C23'] = '=C12*Assumptions!C62'
    ws['D23'] = '=D12*Assumptions!D62'

    ws['A24'] = 'S&M (Sales & Marketing)'
    ws['B24'] = '=B12*Assumptions!B63'
    ws['C24'] = '=C12*Assumptions!C63'
    ws['D24'] = '=D12*Assumptions!D63'

    ws['A25'] = 'G&A (General & Admin)'
    ws['B25'] = '=B12*Assumptions!B64'
    ws['C25'] = '=C12*Assumptions!C64'
    ws['D25'] = '=D12*Assumptions!D64'

    ws['A27'] = 'Total OpEx'
    ws['B27'] = '=SUM(B23:B25)'
    ws['C27'] = '=SUM(C23:C25)'
    ws['D27'] = '=SUM(D23:D25)'
    ws['A27'].font = Font(bold=True)
    ws['A27'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}27'].font = Font(bold=True)
        ws[f'{col}27'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    # EBITDA
    ws['A30'] = 'EBITDA'
    ws['B30'] = '=B18-B27'
    ws['C30'] = '=C18-C27'
    ws['D30'] = '=D18-D27'
    ws['A30'].font = Font(bold=True, size=12)
    ws['A30'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}30'].font = Font(bold=True, size=12)
        ws[f'{col}30'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    ws['A31'] = 'EBITDA Margin'
    ws['B31'] = '=B30/B12'
    ws['C31'] = '=C30/C12'
    ws['D31'] = '=D30/D12'
    for col in ['B', 'C', 'D']:
        ws[f'{col}31'].number_format = '0%'

    # Net Income
    ws['A34'] = 'Depreciation'
    ws['B34'] = 0
    ws['C34'] = 5000
    ws['D34'] = 15000

    ws['A35'] = 'Interest'
    ws['B35'] = 0
    ws['C35'] = 0
    ws['D35'] = 0

    ws['A37'] = 'Net Income'
    ws['B37'] = '=B30-B34-B35'
    ws['C37'] = '=C30-C34-C35'
    ws['D37'] = '=D30-D34-D35'
    ws['A37'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A37'].fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}37'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'{col}37'].fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

    ws['A38'] = 'Net Margin'
    ws['B38'] = '=B37/B12'
    ws['C38'] = '=C37/C12'
    ws['D38'] = '=D37/D12'
    for col in ['B', 'C', 'D']:
        ws[f'{col}38'].number_format = '0%'

    # Formatear todos los números
    for row in range(6, 39):
        for col in ['B', 'C', 'D']:
            if row not in [19, 31, 38]:  # No formatear los % margins
                ws[f'{col}{row}'].number_format = '$#,##0'

    # Ajustar columnas
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 18

def create_cash_flow_advanced(wb):
    """Cash Flow con fórmulas"""
    ws = wb.create_sheet("Cash Flow")

    # Título
    ws['A1'] = 'CASH FLOW STATEMENT'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    ws['B3'] = 'Año 1'
    ws['C3'] = 'Año 2'
    ws['D3'] = 'Año 3'
    for col in ['B', 'C', 'D']:
        ws[f'{col}3'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}3'].fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")

    # OPERATING ACTIVITIES
    ws['A5'] = 'OPERATING ACTIVITIES'
    ws['A5'].font = Font(size=12, bold=True)

    ws['A6'] = 'Net Income'
    ws['B6'] = "='P&L'!B37"
    ws['C6'] = "='P&L'!C37"
    ws['D6'] = "='P&L'!D37"

    ws['A7'] = '+ Depreciation'
    ws['B7'] = "='P&L'!B34"
    ws['C7'] = "='P&L'!C34"
    ws['D7'] = "='P&L'!D34"

    ws['A9'] = 'Cash from Operations'
    ws['B9'] = '=B6+B7'
    ws['C9'] = '=C6+C7'
    ws['D9'] = '=D6+D7'
    ws['A9'].font = Font(bold=True)
    ws['A9'].fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}9'].font = Font(bold=True)
        ws[f'{col}9'].fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")

    # INVESTING ACTIVITIES
    ws['A12'] = 'INVESTING ACTIVITIES'
    ws['A12'].font = Font(size=12, bold=True)

    ws['A13'] = 'CapEx'
    ws['B13'] = -10000
    ws['C13'] = -30000
    ws['D13'] = -100000

    ws['A15'] = 'Cash from Investing'
    ws['B15'] = '=B13'
    ws['C15'] = '=C13'
    ws['D15'] = '=D13'
    ws['A15'].font = Font(bold=True)
    ws['A15'].fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}15'].font = Font(bold=True)
        ws[f'{col}15'].fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")

    # FINANCING ACTIVITIES
    ws['A18'] = 'FINANCING ACTIVITIES'
    ws['A18'].font = Font(size=12, bold=True)

    ws['A19'] = 'Equity Raised'
    ws['B19'] = '=Assumptions!B69'
    ws['C19'] = 0
    ws['D19'] = 0

    ws['A20'] = 'Debt Raised'
    ws['B20'] = 0
    ws['C20'] = 0
    ws['D20'] = 0

    ws['A22'] = 'Cash from Financing'
    ws['B22'] = '=B19+B20'
    ws['C22'] = '=C19+C20'
    ws['D22'] = '=D19+D20'
    ws['A22'].font = Font(bold=True)
    ws['A22'].fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}22'].font = Font(bold=True)
        ws[f'{col}22'].fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")

    # NET CASH FLOW
    ws['A25'] = 'NET CASH FLOW'
    ws['B25'] = '=B9+B15+B22'
    ws['C25'] = '=C9+C15+C22'
    ws['D25'] = '=D9+D15+D22'
    ws['A25'].font = Font(bold=True, size=12)
    ws['A25'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}25'].font = Font(bold=True, size=12)
        ws[f'{col}25'].fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    # CASH BALANCE
    ws['A28'] = 'CASH BALANCE'
    ws['A28'].font = Font(size=12, bold=True)

    ws['A29'] = 'Beginning Cash'
    ws['B29'] = 0
    ws['C29'] = '=B30'
    ws['D29'] = '=C30'

    ws['A30'] = 'Ending Cash'
    ws['B30'] = '=B29+B25'
    ws['C30'] = '=C29+C25'
    ws['D30'] = '=D29+D25'
    ws['A30'].font = Font(bold=True, color="FFFFFF")
    ws['A30'].fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
    for col in ['B', 'C', 'D']:
        ws[f'{col}30'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}30'].fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

    ws['A32'] = 'Monthly Burn Rate'
    ws['B32'] = '=(B9+B15)/12'
    ws['C32'] = '=(C9+C15)/12'
    ws['D32'] = '=(D9+D15)/12'

    ws['A33'] = 'Runway (meses)'
    ws['B33'] = '=IF(B32<0, B30/ABS(B32), "Infinito")'
    ws['C33'] = '=IF(C32<0, C30/ABS(C32), "Infinito")'
    ws['D33'] = '=IF(D32<0, D30/ABS(D32), "Infinito")'

    # Formatear números
    for row in range(6, 34):
        for col in ['B', 'C', 'D']:
            if row != 33:  # No formatear runway
                ws[f'{col}{row}'].number_format = '$#,##0'

    # Ajustar columnas
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 18

def create_scenarios_advanced(wb):
    """Escenarios vinculados a Revenue Model"""
    ws = wb.create_sheet("Scenarios")

    # Título
    ws['A1'] = 'ANÁLISIS DE ESCENARIOS'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Escenarios ARR
    ws['A3'] = 'PROYECCIONES ARR POR ESCENARIO'
    ws['A3'].font = Font(size=12, bold=True)

    headers = ['Año', 'Pesimista (70%)', 'Base (100%)', 'Optimista (150%)', 'Notas']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    ws['A5'] = 'Año 1'
    ws['B5'] = "='Revenue Model'!B43*0.7"
    ws['C5'] = "='Revenue Model'!B43"
    ws['D5'] = "='Revenue Model'!B43*1.5"
    ws['E5'] = 'Setup inicial'

    ws['A6'] = 'Año 2'
    ws['B6'] = "='Revenue Model'!C43*0.7"
    ws['C6'] = "='Revenue Model'!C43"
    ws['D6'] = "='Revenue Model'!C43*1.5"
    ws['E6'] = 'Tracción temprana'

    ws['A7'] = 'Año 3'
    ws['B7'] = "='Revenue Model'!D43*0.7"
    ws['C7'] = "='Revenue Model'!D43"
    ws['D7'] = "='Revenue Model'!D43*1.5"
    ws['E7'] = 'Escala completa'

    # Highlight base case
    for row in [5, 6, 7]:
        ws[f'C{row}'].fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")

    # Formatear
    for row in [5, 6, 7]:
        for col in ['B', 'C', 'D']:
            ws[f'{col}{row}'].number_format = '$#,##0'

    # SENSIBILIDAD
    ws['A10'] = 'ANÁLISIS DE SENSIBILIDAD (Año 3)'
    ws['A10'].font = Font(size=12, bold=True)

    sens_headers = ['Variable', 'Cambio', 'Impacto ARR', 'Nuevo ARR']
    for col, header in enumerate(sens_headers, start=1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    ws['A12'] = 'Base case'
    ws['B12'] = '0%'
    ws['C12'] = '$0'
    ws['D12'] = "='Revenue Model'!D43"

    ws['A13'] = 'Usuarios +10%'
    ws['B13'] = '+10%'
    ws['C13'] = "='Revenue Model'!D43*0.10"
    ws['D13'] = "='Revenue Model'!D43*1.10"

    ws['A14'] = 'Usuarios -10%'
    ws['B14'] = '-10%'
    ws['C14'] = "='Revenue Model'!D43*(-0.10)"
    ws['D14'] = "='Revenue Model'!D43*0.90"

    ws['A15'] = 'Precio servicios +20%'
    ws['B15'] = '+20%'
    ws['C15'] = "='Revenue Model'!D29*0.20"
    ws['D15'] = "='Revenue Model'!D43+C15"

    ws['A16'] = 'Precio servicios -20%'
    ws['B16'] = '-20%'
    ws['C16'] = "='Revenue Model'!D29*(-0.20)"
    ws['D16'] = "='Revenue Model'!D43+C16"

    # Formatear
    for row in range(12, 17):
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'D{row}'].number_format = '$#,##0'

    # RIESGOS
    ws['A20'] = 'PRINCIPALES RIESGOS Y MITIGACIÓN'
    ws['A20'].font = Font(size=12, bold=True)

    risk_headers = ['Riesgo', 'Probabilidad', 'Impacto', 'Mitigación']
    for col, header in enumerate(risk_headers, start=1):
        cell = ws.cell(row=21, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

    risks = [
        ['Adquisición usuarios lenta', 'Media', 'Alto', 'Aumentar S&M budget, partnerships'],
        ['Regulación legal restrictiva', 'Baja', 'Alto', 'Disclaimer claro, no asesoría legal'],
        ['Competencia agresiva', 'Alta', 'Medio', 'Focus en IA diferenciada'],
        ['Costos API Claude suben', 'Media', 'Medio', 'Optimizar prompts, caché, modelos'],
        ['Churn abogados alto', 'Media', 'Alto', 'Programa de éxito, comunidad'],
    ]

    row = 22
    for risk_row in risks:
        for col, value in enumerate(risk_row, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # Ajustar columnas
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 28

def create_dashboard_advanced(wb):
    """Dashboard que resume todo (se crea al final para referenciar otras hojas)"""
    ws = wb.create_sheet("Dashboard", 0)

    # Título
    ws['A1'] = 'JUSTICIAAI - DASHBOARD FINANCIERO'
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # RESUMEN EJECUTIVO
    ws['A3'] = 'RESUMEN EJECUTIVO - MÉTRICAS CLAVE'
    ws['A3'].font = Font(size=14, bold=True)
    ws.merge_cells('A3:H3')

    headers = ['Métrica', 'Año 1', 'Año 2', 'Año 3', 'Growth']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    metrics = [
        ['ARR', "='Revenue Model'!B43", "='Revenue Model'!C43", "='Revenue Model'!D43", "='Revenue Model'!D45"],
        ['MRR', "='Revenue Model'!B44", "='Revenue Model'!C44", "='Revenue Model'!D44", ''],
        ['Usuarios/mes', '=Assumptions!C8', '=Assumptions!D8', '=Assumptions!E8', ''],
        ['Abogados', '=Assumptions!C14', '=Assumptions!D14', '=Assumptions!E14', ''],
        ['Gross Margin', "='P&L'!B19", "='P&L'!C19", "='P&L'!D19", ''],
        ['EBITDA', "='P&L'!B30", "='P&L'!C30", "='P&L'!D30", ''],
        ['EBITDA Margin', "='P&L'!B31", "='P&L'!C31", "='P&L'!D31", ''],
        ['LTV/CAC', "='Unit Economics'!B8", "='Unit Economics'!C8", "='Unit Economics'!D8", ''],
        ['Cash Balance', "='Cash Flow'!B30", "='Cash Flow'!C30", "='Cash Flow'!D30", ''],
    ]

    row = 5
    for metric in metrics:
        ws.cell(row=row, column=1, value=metric[0]).font = Font(bold=True)
        ws.cell(row=row, column=2, value=metric[1])
        ws.cell(row=row, column=3, value=metric[2])
        ws.cell(row=row, column=4, value=metric[3])
        ws.cell(row=row, column=5, value=metric[4])
        row += 1

    # Formatear
    for r in range(5, 14):
        for c in [2, 3, 4]:
            cell = ws.cell(row=r, column=c)
            if r in [5, 6]:  # ARR, MRR
                cell.number_format = '$#,##0'
            elif r in [7, 8]:  # Usuarios, Abogados
                cell.number_format = '#,##0'
            elif r in [9, 11]:  # Margins
                cell.number_format = '0%'
            elif r == 10:  # EBITDA
                cell.number_format = '$#,##0'
            elif r == 12:  # LTV/CAC
                cell.number_format = '0.0"x"'
            elif r == 13:  # Cash
                cell.number_format = '$#,##0'

    ws.cell(row=5, column=5).number_format = '0%'  # Growth

    # REVENUE BREAKDOWN
    ws['A16'] = 'DISTRIBUCIÓN DE INGRESOS - AÑO 3'
    ws['A16'].font = Font(size=12, bold=True)

    rev_headers = ['Stream', 'Monto', '% del Total']
    for col, header in enumerate(rev_headers, start=1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

    ws['A18'] = 'Comisiones'
    ws['B18'] = "='Revenue Model'!D11"
    ws['C18'] = "=B18/'Revenue Model'!D43"

    ws['A19'] = 'Suscripciones'
    ws['B19'] = "='Revenue Model'!D21"
    ws['C19'] = "=B19/'Revenue Model'!D43"

    ws['A20'] = 'Automatizados'
    ws['B20'] = "='Revenue Model'!D29"
    ws['C20'] = "=B20/'Revenue Model'!D43"

    ws['A21'] = 'B2B'
    ws['B21'] = "='Revenue Model'!D35"
    ws['C21'] = "=B21/'Revenue Model'!D43"

    ws['A22'] = 'Partnerships'
    ws['B22'] = "='Revenue Model'!D39"
    ws['C22'] = "=B22/'Revenue Model'!D43"

    ws['A23'] = 'TOTAL'
    ws['B23'] = "='Revenue Model'!D43"
    ws['C23'] = '=SUM(C18:C22)'
    ws['A23'].font = Font(bold=True)
    for col in [1, 2, 3]:
        ws.cell(row=23, column=col).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    # Formatear revenue breakdown
    for r in range(18, 24):
        ws.cell(row=r, column=2).number_format = '$#,##0'
        ws.cell(row=r, column=3).number_format = '0%'

    # NOTA IMPORTANTE
    ws['A26'] = 'MODELO DINÁMICO: Edita los valores en la hoja "Assumptions" y todo se actualiza automáticamente'
    ws['A26'].font = Font(bold=True, italic=True, color="7C3AED")
    ws['A26'].fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    ws.merge_cells('A26:H26')
    ws['A26'].alignment = Alignment(horizontal='center')

    # Ajustar columnas
    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 15

if __name__ == "__main__":
    print("🚀 Generando modelo financiero AVANZADO de JusticiaAI...")
    print("🔗 Con fórmulas dinámicas entre todas las hojas...")
    create_financial_model_advanced()
    print("✅ ¡Listo! Abre el archivo: JusticiaAI-Financial-Model-Advanced.xlsx")
    print("💡 Tip: Cambia valores en la hoja 'Assumptions' y verás cómo se actualiza todo automáticamente")
